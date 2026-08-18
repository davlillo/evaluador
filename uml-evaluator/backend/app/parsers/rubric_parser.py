"""
Parser de la rúbrica Excel (.xlsx) de cantidades esperadas por tipo de
diagrama. Produce un EvaluationProfile por tipo ('class', 'usecase',
'sequence'), listo para pasar a UMLComparator. Ver
app/parsers/rubric_schema.py para el esquema de hojas/columnas/alias, y
app/parsers/rubric_template_builder.py para la plantilla descargable.
"""
import unicodedata
from typing import Any, Dict, List

from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet

from app.comparator.scoring_modes import (
    ClassRubricRule, EvaluationProfile, ExpectedCount, ScoringMode,
)
from app.parsers.rubric_schema import (
    SHEET_ALIASES, SHEET_CLASS_RUBRIC, SHEET_CONFIG,
    SHEET_TO_DIAGRAM_TYPE, VALID_MODES, MODE_LABEL_TO_INTERNAL,
    RETIRED_MODE_ALIASES,
)


class RubricParseError(Exception):
    """Errores de formato acumulados durante el parseo, con ubicación por celda."""

    def __init__(self, errors: List[str]):
        self.errors = errors
        super().__init__("; ".join(errors))


def _normalized_option(value: Any) -> str:
    text = unicodedata.normalize("NFD", str(value or "").strip().lower())
    return "".join(char for char in text if unicodedata.category(char) != "Mn").replace(" ", "_")


def _cell_ref(sheet_name: str, row: int, col_letter: str) -> str:
    return f"{sheet_name}!{col_letter}{row}"


def _resolve_mode(raw_value: str) -> str | None:
    """Traduce lo que haya en la celda de modo al código interno.

    Acepta, en este orden: la etiqueta en español de la plantilla vigente,
    el código interno, y las etiquetas/códigos de modos ya retirados (que
    se mapean al equivalente vigente para no rechazar rúbricas viejas).
    """
    if raw_value in MODE_LABEL_TO_INTERNAL:
        return MODE_LABEL_TO_INTERNAL[raw_value]
    if raw_value in VALID_MODES:
        return raw_value
    return RETIRED_MODE_ALIASES.get(raw_value)


def _parse_config_sheet(ws: Worksheet, errors: List[str]) -> Dict[str, Any]:
    config: Dict[str, Any] = {"mode": "expected_with_penalty"}

    for row_idx in range(2, ws.max_row + 1):
        label = ws.cell(row=row_idx, column=1).value
        value = ws.cell(row=row_idx, column=2).value
        if label is None:
            break
        if str(label).strip() != "Modo de evaluación" or value is None:
            continue

        mode_str = str(value).strip()
        internal_mode = _resolve_mode(mode_str)
        if internal_mode is None:
            valid_labels = ', '.join(MODE_LABEL_TO_INTERNAL.keys())
            errors.append(
                f"{_cell_ref(SHEET_CONFIG, row_idx, 'B')}: modo '{mode_str}' no reconocido. "
                f"Elegí una opción de la lista desplegable: {valid_labels}."
            )
            continue
        config["mode"] = internal_mode

    return config


def _parse_element_sheet(
    ws: Worksheet, sheet_name: str, errors: List[str],
) -> Dict[str, ExpectedCount]:
    alias_map = dict(SHEET_ALIASES[sheet_name])
    expected_counts: Dict[str, ExpectedCount] = {}

    # Las filas de datos son un bloque contiguo desde la fila 2 (la
    # plantilla puede llevar una nota informativa más abajo, separada por
    # una fila en blanco; esa nota no es una fila de datos).
    last_data_row = 1
    for row_idx in range(2, ws.max_row + 1):
        if ws.cell(row=row_idx, column=1).value is None:
            break
        last_data_row = row_idx

    for row_idx in range(2, last_data_row + 1):
        label = ws.cell(row=row_idx, column=1).value
        if label is None or str(label).strip() == "":
            continue
        label = str(label).strip()
        element_type = alias_map.get(label)
        if element_type is None:
            errors.append(
                f"{_cell_ref(sheet_name, row_idx, 'A')}: elemento '{label}' no reconocido para esta hoja."
            )
            continue

        qty_raw = ws.cell(row=row_idx, column=2).value
        if qty_raw is None:
            continue

        try:
            qty = int(qty_raw)
            if qty < 0:
                raise ValueError
        except (TypeError, ValueError):
            errors.append(
                f"{_cell_ref(sheet_name, row_idx, 'B')}: 'Cantidad esperada' debe ser "
                f"un entero no negativo (recibido {qty_raw!r})."
            )
            continue

        expected_counts[element_type] = ExpectedCount(
            element_type=element_type, expected_quantity=qty, label=label,
        )

    return expected_counts


def _parse_class_rubric_sheet(
    ws: Worksheet,
    errors: List[str],
) -> List[ClassRubricRule]:
    rules: List[ClassRubricRule] = []
    valid_types = {"classes", "relationship", "multiplicity", "association_class"}

    for row_idx in range(2, ws.max_row + 1):
        raw_type = ws.cell(row=row_idx, column=1).value
        if raw_type is None:
            break
        criterion_type = {
            "clases": "classes",
            "cantidad_de_clases": "classes",
            "relacion": "relationship",
            "multiplicidad": "multiplicity",
            "clase_de_asociacion": "association_class",
        }.get(_normalized_option(raw_type), _normalized_option(raw_type))
        if criterion_type not in valid_types:
            errors.append(
                f"{_cell_ref(SHEET_CLASS_RUBRIC, row_idx, 'A')}: tipo "
                f"'{criterion_type}' no reconocido."
            )
            continue

        label = str(ws.cell(row=row_idx, column=2).value or "").strip()
        try:
            weight = float(ws.cell(row=row_idx, column=3).value)
            if weight < 0 or weight > 100:
                raise ValueError
        except (TypeError, ValueError):
            errors.append(
                f"{_cell_ref(SHEET_CLASS_RUBRIC, row_idx, 'C')}: el peso "
                "debe estar entre 0 y 100."
            )
            continue

        source = str(ws.cell(row=row_idx, column=4).value or "").strip() or None
        target = str(ws.cell(row=row_idx, column=5).value or "").strip() or None
        relationship_type_raw = str(
            ws.cell(row=row_idx, column=6).value
            or ("association_class" if criterion_type == "association_class" else "association")
        ).strip()
        relationship_type = {
            "asociacion": "association",
            "agregacion": "aggregation",
            "composicion": "composition",
            "clase_de_asociacion": "association_class",
        }.get(_normalized_option(relationship_type_raw), _normalized_option(relationship_type_raw))
        multiplicity_end_raw = ws.cell(row=row_idx, column=7).value
        multiplicity_end = {
            "origen": "source",
            "destino": "target",
        }.get(_normalized_option(multiplicity_end_raw), _normalized_option(multiplicity_end_raw)) or None
        expected_value = str(ws.cell(row=row_idx, column=8).value or "").strip() or None
        quantity_raw = ws.cell(row=row_idx, column=9).value
        expected_quantity = None
        if quantity_raw not in (None, ""):
            try:
                expected_quantity = int(quantity_raw)
                if expected_quantity < 0:
                    raise ValueError
            except (TypeError, ValueError):
                errors.append(
                    f"{_cell_ref(SHEET_CLASS_RUBRIC, row_idx, 'I')}: la "
                    "cantidad esperada debe ser un entero no negativo."
                )
                continue

        if criterion_type != "classes" and (not source or not target):
            errors.append(
                f"{_cell_ref(SHEET_CLASS_RUBRIC, row_idx, 'D')}: origen y "
                "destino son obligatorios para criterios de relación."
            )
            continue
        if criterion_type == "multiplicity" and (
            multiplicity_end not in {"source", "target"} or expected_value is None
        ):
            errors.append(
                f"{_cell_ref(SHEET_CLASS_RUBRIC, row_idx, 'G')}: indicá "
                "source/target y la multiplicidad esperada."
            )
            continue

        rules.append(ClassRubricRule(
            rule_id=f"xlsx-{row_idx}",
            criterion_type=criterion_type,
            label=label or f"Criterio {row_idx}",
            weight=weight,
            expected_quantity=expected_quantity,
            source=source,
            target=target,
            relationship_type=relationship_type,
            multiplicity_end=multiplicity_end,
            expected_multiplicity=expected_value,
        ))

    total = sum(rule.weight for rule in rules)
    if rules and abs(total - 100.0) >= 0.01:
        errors.append(
            f"{SHEET_CLASS_RUBRIC}: los pesos suman {total:g}% y deben sumar 100%."
        )
    return rules


def parse_rubric_workbook(wb) -> Dict[str, EvaluationProfile]:
    """Parsea un Workbook ya cargado (openpyxl) y retorna un EvaluationProfile
    por tipo de diagrama detectado. Lanza RubricParseError si hay celdas
    inválidas (agrupa todos los errores encontrados, no solo el primero)."""
    errors: List[str] = []

    config: Dict[str, Any] = {}
    if SHEET_CONFIG in wb.sheetnames:
        config = _parse_config_sheet(wb[SHEET_CONFIG], errors)
    else:
        errors.append(f"Falta la hoja '{SHEET_CONFIG}'.")

    profiles: Dict[str, EvaluationProfile] = {}
    for sheet_name, diagram_type in SHEET_TO_DIAGRAM_TYPE.items():
        if sheet_name not in wb.sheetnames:
            continue
        expected_counts = _parse_element_sheet(wb[sheet_name], sheet_name, errors)
        if not expected_counts:
            continue

        profiles[diagram_type] = EvaluationProfile(
            mode=ScoringMode(config.get("mode", "expected_with_penalty")),
            expected_counts=expected_counts,
        )

    if SHEET_CLASS_RUBRIC in wb.sheetnames:
        class_rules = _parse_class_rubric_sheet(wb[SHEET_CLASS_RUBRIC], errors)
        if class_rules:
            class_profile = profiles.get("class") or EvaluationProfile(
                mode=ScoringMode.EXPECTED_WITH_PENALTY,
            )
            class_profile.class_rules = class_rules
            profiles["class"] = class_profile

    if not profiles and not errors:
        errors.append("No se encontraron hojas reconocidas (Clases, CasosDeUso, Secuencia).")

    if errors:
        raise RubricParseError(errors)

    return profiles


def parse_rubric_xlsx(file_path: str) -> Dict[str, EvaluationProfile]:
    """Lee un archivo .xlsx de rúbrica y retorna un EvaluationProfile por
    tipo de diagrama. Ver parse_rubric_workbook para el detalle."""
    wb = load_workbook(file_path, data_only=True)
    return parse_rubric_workbook(wb)
