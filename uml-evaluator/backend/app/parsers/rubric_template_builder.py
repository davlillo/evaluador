"""
Construye la plantilla Excel de rúbrica de evaluación (cantidades esperadas
por tipo de diagrama) que el profesor descarga desde GET /api/rubric-template,
llena, y vuelve a subir vía POST /api/rubric/parse. Ver rubric_schema.py para
el esquema de hojas/columnas/alias compartido con rubric_parser.py.
"""
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.worksheet.datavalidation import DataValidation

from app.parsers.rubric_schema import (
    CLASS_RUBRIC_COLUMNS, COLUMNS, SHEET_ALIASES, SHEET_CLASS,
    SHEET_CLASS_RUBRIC, SHEET_USECASE, SHEET_SEQUENCE, SHEET_CONFIG,
    CONFIG_ROWS, VALID_MODE_LABELS,
)

SHEET_MODE_OPTIONS = "_ModosDisponibles"

HEADER_FILL = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
HEADER_FONT = Font(color="FFFFFF", bold=True)
TITLE_FONT = Font(bold=True, size=13)
NOTE_FONT = Font(italic=True, size=9, color="666666")

EXAMPLE_VALUES = {
    "class": {"Clases": 4, "Atributos totales": 12, "Métodos totales": 8, "Asociación": 3,
              "Agregación": 2, "Composición": 0, "Herencia": 1, "Implementación": 0},
    "usecase": {"Actores": 2, "Casos de uso": 4, "Relaciones actor-CU": 4,
                "Include": 2, "Extend": 1},
    "sequence": {"Líneas de vida": 3, "Mensajes síncronos": 5, "Mensajes asíncronos": 1,
                 "Mensajes de creación": 1, "Fragmentos alt": 1, "Fragmentos loop": 1},
}

SHEET_DIAGRAM_KIND = {SHEET_CLASS: "class", SHEET_USECASE: "usecase", SHEET_SEQUENCE: "sequence"}


def _write_sheet(wb: Workbook, sheet_name: str) -> None:
    ws = wb.create_sheet(sheet_name)
    ws.append(COLUMNS)
    for col_idx in range(1, len(COLUMNS) + 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center")

    aliases = SHEET_ALIASES[sheet_name]
    examples = EXAMPLE_VALUES[SHEET_DIAGRAM_KIND[sheet_name]]
    for row_idx, (label, _element_type) in enumerate(aliases, start=2):
        ws.cell(row=row_idx, column=1, value=label)
        ws.cell(row=row_idx, column=2, value=examples.get(label, 0))
        ws.cell(row=row_idx, column=5, value="")

    # Lista desplegable para la columna "Elemento" (evita typos).
    allowed_labels = [label for label, _ in aliases]
    dv = DataValidation(
        type="list",
        formula1='"' + ",".join(allowed_labels) + '"',
        allow_blank=True,
    )
    ws.add_data_validation(dv)
    dv.add(f"A2:A{len(aliases) + 1}")

    ws.column_dimensions["A"].width = 26
    ws.column_dimensions["B"].width = 18
    ws.column_dimensions["C"].width = 45

    note = ws.cell(row=len(aliases) + 3, column=1,
                    value="Escribí en 'Cantidad esperada' cuántos elementos de cada tipo esperás "
                          "que tenga el diagrama. La columna 'Notas' es libre y no afecta la calificación.")
    note.font = NOTE_FONT
    note.alignment = Alignment(wrap_text=True)
    ws.row_dimensions[len(aliases) + 3].height = 28
    ws.merge_cells(start_row=len(aliases) + 3, start_column=1, end_row=len(aliases) + 3, end_column=3)


def _write_class_rubric_sheet(wb: Workbook) -> None:
    ws = wb.create_sheet(SHEET_CLASS_RUBRIC)
    ws.append(CLASS_RUBRIC_COLUMNS)
    for cell in ws[1]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center")

    rows = [
        ["Clases", "Clases", 20, "", "", "", "", "", 5],
        ["Multiplicidad", "Multiplicidad 1 en Afiliado", 10, "Afiliado", "Ganado", "Asociación", "Origen", "1", ""],
        ["Multiplicidad", "Multiplicidad 1..* en Ganado", 10, "Afiliado", "Ganado", "Asociación", "Destino", "1..*", ""],
        ["Multiplicidad", "Multiplicidad 1 en Ganado", 10, "Ganado", "Producción", "Asociación", "Origen", "1", ""],
        ["Multiplicidad", "Multiplicidad 1..* en Producción", 10, "Ganado", "Producción", "Asociación", "Destino", "1..*", ""],
        ["Clase de asociación", "Clase de asociación Enfermedad-Ganado", 20, "Enfermedad", "Ganado", "Clase de asociación", "", "", ""],
        ["Multiplicidad", "Multiplicidad 0..* en Tratamiento", 10, "Tratamiento", "Medicamento", "Agregación", "Origen", "0..*", ""],
        ["Multiplicidad", "Multiplicidad 1..* en Medicamento", 10, "Tratamiento", "Medicamento", "Agregación", "Destino", "1..*", ""],
    ]
    for row in rows:
        ws.append(row)

    widths = [20, 44, 12, 20, 20, 20, 12, 20, 20]
    for index, width in enumerate(widths, start=1):
        ws.column_dimensions[chr(64 + index)].width = width
    ws.freeze_panes = "A2"
    validations = [
        (DataValidation(type="list", formula1='"Clases,Multiplicidad,Clase de asociación"'), "A2:A200"),
        (DataValidation(type="list", formula1='"Asociación,Agregación,Composición,Clase de asociación"'), "F2:F200"),
        (DataValidation(type="list", formula1='"Origen,Destino"'), "G2:G200"),
        (DataValidation(type="list", formula1='"0,1,0..1,1..1,*,0..*,1..*"'), "H2:H200"),
    ]
    for validation, cell_range in validations:
        validation.allow_blank = True
        ws.add_data_validation(validation)
        validation.add(cell_range)

    note_row = len(rows) + 3
    ws.cell(
        row=note_row,
        column=1,
        value=(
            "Los pesos deben sumar 100%. Cada fila de multiplicidad se califica "
            "independientemente. Usá las listas desplegables para seleccionar "
            "asociación, agregación o composición y el extremo correspondiente."
        ),
    )
    ws.cell(row=note_row, column=1).font = NOTE_FONT
    ws.cell(row=note_row, column=1).alignment = Alignment(wrap_text=True)
    ws.merge_cells(start_row=note_row, start_column=1, end_row=note_row, end_column=9)


def _write_mode_options_sheet(wb: Workbook) -> str:
    """Hoja auxiliar oculta con las etiquetas de modo en español, usada como
    fuente de la lista desplegable (evita el límite de 255 caracteres de
    Excel para listas inline con textos largos)."""
    ws = wb.create_sheet(SHEET_MODE_OPTIONS)
    for row_idx, label in enumerate(VALID_MODE_LABELS, start=1):
        ws.cell(row=row_idx, column=1, value=label)
    ws.sheet_state = "hidden"
    return f"{SHEET_MODE_OPTIONS}!$A$1:$A${len(VALID_MODE_LABELS)}"


def _write_config_sheet(wb: Workbook, mode_options_range: str) -> None:
    ws = wb.create_sheet(SHEET_CONFIG)
    ws.append(["Parámetro", "Valor"])
    for col_idx in (1, 2):
        cell = ws.cell(row=1, column=col_idx)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT

    for row_idx, (label, default) in enumerate(CONFIG_ROWS, start=2):
        ws.cell(row=row_idx, column=1, value=label)
        ws.cell(row=row_idx, column=2, value=default)

    dv = DataValidation(type="list", formula1=f"={mode_options_range}", allow_blank=False)
    ws.add_data_validation(dv)
    dv.add("B2")

    ws.column_dimensions["A"].width = 38
    ws.column_dimensions["B"].width = 45

    note = ws.cell(row=len(CONFIG_ROWS) + 3, column=1, value=(
        "El 'Modo de evaluación' define CÓMO se calcula la nota (no es obligatorio cambiarlo: "
        "el valor por defecto ya es válido para presentar). Elegí una opción de la lista "
        "desplegable en la celda B2 — no escribas texto libre."
    ))
    note.font = NOTE_FONT
    note.alignment = Alignment(wrap_text=True)
    ws.row_dimensions[len(CONFIG_ROWS) + 3].height = 45
    ws.merge_cells(
        start_row=len(CONFIG_ROWS) + 3, start_column=1,
        end_row=len(CONFIG_ROWS) + 3, end_column=2,
    )

    explanations = [
        "Similitud (comparar contra la solución): compara el diagrama del estudiante contra el modelo de referencia. Es el modo tradicional.",
        "Cantidades esperadas - sin descuento: solo importa cumplir el mínimo (columna 'Cantidad esperada' de cada hoja). Entregar de más no resta puntos.",
        "Cantidades esperadas - con descuento: igual al anterior, pero entregar de más SÍ resta puntos (según 'Descuento por unidad de exceso').",
        "Similitud con descuento: como el primer modo, pero resta puntos extra si el estudiante entrega más elementos que la solución.",
        "Híbrido: combina similitud y cantidades esperadas, según 'Peso similitud'.",
        "Rango aceptable: usa las columnas Mínimo/Máximo de cada hoja en vez de 'Cantidad esperada'.",
    ]
    start_row = len(CONFIG_ROWS) + 5
    ws.cell(row=start_row, column=1, value="Qué significa cada modo:").font = Font(bold=True, size=10)
    for i, text in enumerate(explanations):
        row = start_row + 1 + i
        cell = ws.cell(row=row, column=1, value=text)
        cell.font = NOTE_FONT
        cell.alignment = Alignment(wrap_text=True)
        ws.row_dimensions[row].height = 28
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=2)


def generate_rubric_template(output_path: Path) -> Path:
    """Genera la plantilla Excel de rúbrica en output_path y retorna la ruta."""
    wb = Workbook()
    wb.remove(wb.active)  # quitar la hoja "Sheet" por defecto

    intro = wb.create_sheet("Instrucciones")
    intro["A1"] = "Plantilla de rúbrica de evaluación — UML Evaluator"
    intro["A1"].font = TITLE_FONT
    intro["A3"] = (
        "Para diagramas de clases, complete RubricaClases con cada multiplicidad "
        "y clase de asociación que desea evaluar; sus pesos deben sumar 100%. "
        "CasosDeUso y Secuencia mantienen sus hojas de cantidades esperadas."
    )
    intro["A3"].alignment = Alignment(wrap_text=True)
    intro.column_dimensions["A"].width = 100
    intro.row_dimensions[3].height = 60

    _write_sheet(wb, SHEET_CLASS)
    _write_class_rubric_sheet(wb)
    # La hoja antigua se conserva oculta para compatibilidad con rúbricas ya
    # distribuidas; la nueva evaluación de clases no usa atributos ni métodos.
    wb[SHEET_CLASS].sheet_state = "hidden"
    wb.active = wb.sheetnames.index(SHEET_CLASS_RUBRIC)
    _write_sheet(wb, SHEET_USECASE)
    _write_sheet(wb, SHEET_SEQUENCE)
    mode_options_range = _write_mode_options_sheet(wb)
    _write_config_sheet(wb, mode_options_range)

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)
    return output_path
