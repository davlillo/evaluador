"""
Exporta el resultado de una evaluación de lote a un archivo Excel (.xlsx)
con 3 hojas: Notas (resumen por estudiante), Detalle (criterio por
criterio, con las cantidades esperada/registrada y el factor de curva
aplicado), y Resumen (metadatos del lote).

Reemplaza el CSV plano anterior (app/src/lib/batch-csv.ts en el frontend),
que solo tenía Carné/Similitud/Nota/Estado sin ningún desglose.

El frontend reenvía el JSON del batch tal cual lo tiene en pantalla (no se
re-evalúa acá), junto con las notas que el docente haya editado a mano
(nota_overrides), porque esas ediciones solo viven en el cliente.
"""
import io
from datetime import datetime
from typing import Any, Dict, Optional

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

from app.grading import percent_to_nota, is_aprobado

HEADER_FILL = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
HEADER_FONT = Font(color="FFFFFF", bold=True)
TITLE_FONT = Font(bold=True, size=13)

DIAGRAM_LABELS = {"class": "Clases", "usecase": "Casos de uso", "sequence": "Secuencia"}
CRITERION_LABELS = {
    "classes": "Cantidad de clases",
    "relationship": "Relación",
    "multiplicity": "Multiplicidad",
    "association_class": "Clase de asociación",
}
RELATIONSHIP_LABELS = {
    "association": "Asociación",
    "aggregation": "Agregación",
    "composition": "Composición",
    "association_class": "Clase de asociación",
}


def _write_header(ws, headers: list[str]) -> None:
    ws.append(headers)
    for col_idx in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}1"


def _run_similarity(run: Dict[str, Any], diagram_type: str) -> Optional[float]:
    if run.get("status") != "ok":
        return None
    comparison = run.get("comparison") or {}
    return comparison.get("overall_similarity", run.get("similarity"))


def _write_notas_sheet(wb: Workbook, batch: Dict[str, Any], nota_overrides: Dict[str, float]) -> None:
    ws = wb.create_sheet("Notas")
    diagram_types = [d for d in ("class", "usecase", "sequence") if d in DIAGRAM_LABELS]
    headers = ["Carné", "Similitud (%)", "Nota (0-10)", "Estado"] + [
        f"{DIAGRAM_LABELS[d]} (%)" for d in diagram_types
    ]
    _write_header(ws, headers)

    for r in batch.get("results", []):
        student_id = r.get("student_id", "")
        if r.get("status") == "error":
            ws.append([student_id, None, None, "Error"] + [None] * len(diagram_types))
            continue

        final_score = float(r.get("final_score", 0.0))
        nota = nota_overrides.get(student_id)
        if nota is None:
            nota = r.get("nota", percent_to_nota(final_score))
        estado = "Completo" if r.get("complete") else "Incompleto"

        runs = r.get("runs", {})
        row = [student_id, round(final_score, 2), round(float(nota), 1), estado]
        for d in diagram_types:
            sim = _run_similarity(runs.get(d, {}), d)
            row.append(round(sim, 2) if sim is not None else None)
        ws.append(row)

    ws.column_dimensions["A"].width = 16
    for col in ("B", "C", "D"):
        ws.column_dimensions[col].width = 14
    for i in range(len(diagram_types)):
        ws.column_dimensions[get_column_letter(5 + i)].width = 16


def _iter_criteria(comparison: Dict[str, Any]):
    """Recorre los criterios de un ComparisonResult ya serializado (to_dict),
    combinando breakdown (expected/found/correct/similarity) con
    penalty_breakdown (factor/expected_used/delivered/penalty_applied)
    cuando exista."""
    class_rubric = comparison.get("class_rubric_breakdown") or []
    if class_rubric:
        for row in class_rubric:
            expected = row.get("expected", "")
            modeled = row.get("modeled", "")
            yield (
                row.get("label", row.get("rule_id", "criterio")),
                expected,
                modeled,
                None,
                row.get("score", 0.0),
                row.get("criterion_type", ""),
                row.get("relationship_type", ""),
                row.get("modeled_relationship_type", ""),
                row.get("source", ""),
                row.get("target", ""),
                row.get("multiplicity_end", ""),
                row.get("weight", 0.0),
                row.get("contribution", 0.0),
                row.get("message", ""),
            )
        return

    breakdown = comparison.get("breakdown") or {}
    penalty_breakdown = comparison.get("penalty_breakdown") or {}

    for key, slice_data in breakdown.items():
        if not isinstance(slice_data, dict) or "similarity" not in slice_data:
            continue
        penalty = penalty_breakdown.get(key, {})
        expected = penalty.get("expected_used", slice_data.get("expected", 0))
        delivered = penalty.get("delivered", slice_data.get("found", 0))
        factor = penalty.get("factor")
        score = penalty.get("score", slice_data.get("similarity", 0.0))
        yield key, expected, delivered, factor, score, "", "", "", "", "", "", "", "", ""


def _write_detalle_sheet(wb: Workbook, batch: Dict[str, Any]) -> None:
    ws = wb.create_sheet("Detalle")
    headers = [
        "Carné", "Diagrama", "Criterio", "Cantidad esperada", "Cantidad ingresada",
        "Diferencia", "Factor de penalización aplicado", "Puntaje obtenido",
        "Tipo de criterio", "Tipo configurado", "Tipo detectado", "Clase origen", "Clase destino",
        "Extremo evaluado", "Peso (%)", "Aporte ponderado (%)", "Explicación",
    ]
    _write_header(ws, headers)

    for r in batch.get("results", []):
        if r.get("status") == "error":
            continue
        student_id = r.get("student_id", "")
        for diagram_type, run in (r.get("runs") or {}).items():
            if run.get("status") != "ok":
                continue
            comparison = run.get("comparison") or {}
            for (
                key, expected, delivered, factor, score, criterion_type,
                relationship_type, modeled_relationship_type, source, target,
                end, weight, contribution, message,
            ) in _iter_criteria(comparison):
                diferencia = (
                    delivered - expected
                    if isinstance(delivered, (int, float))
                    and isinstance(expected, (int, float))
                    else "—"
                )
                ws.append([
                    student_id,
                    DIAGRAM_LABELS.get(diagram_type, diagram_type),
                    key,
                    expected,
                    delivered,
                    diferencia,
                    round(factor, 4) if factor is not None else "—",
                    round(score, 2) if score is not None else None,
                    CRITERION_LABELS.get(criterion_type, criterion_type or "—"),
                    RELATIONSHIP_LABELS.get(relationship_type, relationship_type or "—"),
                    RELATIONSHIP_LABELS.get(
                        modeled_relationship_type,
                        modeled_relationship_type or "—",
                    ),
                    source or "—",
                    target or "—",
                    {"source": "Origen", "target": "Destino"}.get(end, end or "—"),
                    round(weight, 2) if isinstance(weight, (int, float)) else weight,
                    round(contribution, 2) if isinstance(contribution, (int, float)) else contribution,
                    message or "—",
                ])

    widths = [16, 14, 28, 18, 22, 12, 18, 16, 18, 18, 18, 22, 22, 18, 12, 18, 52]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w


def _write_resumen_sheet(wb: Workbook, batch: Dict[str, Any]) -> None:
    ws = wb.create_sheet("Resumen")
    ws.append(["Parámetro", "Valor"])
    for col_idx in (1, 2):
        cell = ws.cell(row=1, column=col_idx)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT

    results = batch.get("results", [])
    completos = sum(1 for r in results if r.get("complete"))
    scores = [float(r.get("final_score", 0.0)) for r in results if r.get("status") != "error"]
    aprobados = sum(
        1 for r in results
        if r.get("status") != "error" and is_aprobado(r.get("nota", percent_to_nota(r.get("final_score", 0.0))))
    )
    promedio = round(sum(scores) / len(scores), 2) if scores else 0.0

    global_weights = batch.get("global_weights_used", {})
    rows = [
        ("Fecha de exportación", datetime.now().strftime("%Y-%m-%d %H:%M")),
        ("Estudiantes evaluados", len(results)),
        ("Evaluaciones completas", completos),
        ("Evaluaciones incompletas", len(results) - completos),
        ("Promedio de similitud (%)", promedio),
        ("Aprobados", aprobados),
        ("Reprobados", len(results) - aprobados),
        ("Peso global — Clases (%)", global_weights.get("class")),
        ("Peso global — Casos de uso (%)", global_weights.get("usecase")),
        ("Peso global — Secuencia (%)", global_weights.get("sequence")),
    ]
    for label, value in rows:
        ws.append([label, value])

    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 20


def build_batch_xlsx(batch: Dict[str, Any], nota_overrides: Optional[Dict[str, float]] = None) -> bytes:
    """Construye el .xlsx del lote y retorna los bytes listos para servir."""
    wb = Workbook()
    wb.remove(wb.active)

    overrides = nota_overrides or {}
    _write_notas_sheet(wb, batch, overrides)
    _write_detalle_sheet(wb, batch)
    _write_resumen_sheet(wb, batch)

    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()
