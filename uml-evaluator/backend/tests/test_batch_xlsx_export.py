"""
Tests del export a Excel de un lote (app.exporters.batch_xlsx), que
reemplaza al CSV plano anterior con 3 hojas: Notas, Detalle y Resumen.
"""
import io

import pytest
from openpyxl import load_workbook

from app.exporters.batch_xlsx import build_batch_xlsx


def _sample_batch():
    return {
        "results": [
            {
                "student_id": "AB12345",
                "status": "ok",
                "complete": True,
                "final_score": 88.5,
                "nota": 8.9,
                "runs": {
                    "class": {
                        "status": "ok",
                        "similarity": 90.0,
                        "comparison": {
                            "overall_similarity": 90.0,
                            "scoring_mode": "expected_with_penalty",
                            "breakdown": {
                                "classes": {"similarity": 66.67, "expected": 4, "found": 6, "correct": 4},
                            },
                            "penalty_breakdown": {
                                "classes": {
                                    "score": 66.67, "base_score": 100.0, "penalty_applied": 33.33,
                                    "expected_used": 4, "delivered": 6, "factor": 0.6667,
                                    "explanation": "curva",
                                },
                            },
                        },
                    },
                },
            },
            {
                "student_id": "CD67890",
                "status": "error",
                "complete": False,
                "final_score": 0.0,
                "error": "No se pudo parsear",
                "runs": {},
            },
        ],
        "global_weights_used": {"class": 40.0, "usecase": 35.0, "sequence": 25.0},
    }


class TestHojasPresentes:
    def test_genera_las_tres_hojas(self):
        content = build_batch_xlsx(_sample_batch())
        wb = load_workbook(io.BytesIO(content))
        assert wb.sheetnames == ["Notas", "Detalle", "Resumen"]


class TestHojaNotas:
    def test_columnas_y_filas(self):
        content = build_batch_xlsx(_sample_batch())
        wb = load_workbook(io.BytesIO(content), data_only=True)
        ws = wb["Notas"]
        header = [c.value for c in ws[1]]
        assert header[:4] == ["Carné", "Similitud (%)", "Nota (0-10)", "Estado"]

        row_ab = [c.value for c in ws[2]]
        assert row_ab[0] == "AB12345"
        assert row_ab[2] == 8.9
        assert row_ab[3] == "Completo"

        row_cd = [c.value for c in ws[3]]
        assert row_cd[0] == "CD67890"
        assert row_cd[3] == "Error"

    def test_override_de_nota_gana_sobre_la_del_backend(self):
        content = build_batch_xlsx(_sample_batch(), nota_overrides={"AB12345": 10.0})
        wb = load_workbook(io.BytesIO(content), data_only=True)
        ws = wb["Notas"]
        row_ab = [c.value for c in ws[2]]
        assert row_ab[2] == 10.0


class TestHojaDetalle:
    def test_columnas_pedidas_explicitamente(self):
        content = build_batch_xlsx(_sample_batch())
        wb = load_workbook(io.BytesIO(content), data_only=True)
        ws = wb["Detalle"]
        header = [c.value for c in ws[1]]
        assert header[:8] == [
            "Carné", "Diagrama", "Criterio", "Cantidad esperada", "Cantidad ingresada",
            "Diferencia", "Factor de penalización aplicado", "Puntaje obtenido",
        ]
        assert header[8:] == [
            "Tipo de criterio", "Tipo configurado", "Tipo detectado", "Clase origen", "Clase destino",
            "Extremo evaluado", "Peso (%)", "Aporte ponderado (%)", "Explicación",
        ]

    def test_fila_de_criterio_con_exceso(self):
        content = build_batch_xlsx(_sample_batch())
        wb = load_workbook(io.BytesIO(content), data_only=True)
        ws = wb["Detalle"]
        row = [c.value for c in ws[2]]
        assert row[0] == "AB12345"
        assert row[3] == 4       # esperada
        assert row[4] == 6       # ingresada
        assert row[5] == 2       # diferencia (positiva = exceso)
        assert row[6] == pytest.approx(0.6667, abs=0.001)
        assert row[7] == pytest.approx(66.67, abs=0.01)

    def test_estudiante_con_error_no_genera_filas(self):
        content = build_batch_xlsx(_sample_batch())
        wb = load_workbook(io.BytesIO(content), data_only=True)
        ws = wb["Detalle"]
        student_ids = [c.value for c in ws["A"] if c.row > 1]
        assert "CD67890" not in student_ids


class TestHojaResumen:
    def test_contiene_pesos_globales(self):
        content = build_batch_xlsx(_sample_batch())
        wb = load_workbook(io.BytesIO(content), data_only=True)
        ws = wb["Resumen"]
        rows = {row[0].value: row[1].value for row in ws.iter_rows(min_row=2) if row[0].value}
        assert rows["Peso global — Clases (%)"] == 40.0
        assert rows["Estudiantes evaluados"] == 2
