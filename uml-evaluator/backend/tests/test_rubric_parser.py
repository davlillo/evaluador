"""
Tests del parser de rúbrica Excel (app.parsers.rubric_parser), usando la
plantilla generada por app.parsers.rubric_template_builder como base y
mutándola con openpyxl para simular lo que un profesor completaría.
"""
import pytest
from openpyxl import load_workbook

from app.comparator.scoring_modes import ScoringMode
from app.parsers.rubric_parser import parse_rubric_workbook, RubricParseError
from app.parsers.rubric_template_builder import generate_rubric_template


@pytest.fixture
def template_path(tmp_path):
    return generate_rubric_template(tmp_path / "rubric.xlsx")


class TestPlantillaGenerada:
    def test_plantilla_por_defecto_parsea_sin_errores(self, template_path):
        wb = load_workbook(template_path, data_only=True)
        profiles = parse_rubric_workbook(wb)
        assert set(profiles.keys()) == {"class", "usecase", "sequence"}

    def test_valores_de_ejemplo_de_clases(self, template_path):
        wb = load_workbook(template_path, data_only=True)
        profiles = parse_rubric_workbook(wb)
        class_profile = profiles["class"]
        assert class_profile.expected_counts["classes"].expected_quantity == 4
        assert class_profile.expected_counts["aggregation"].expected_quantity == 2
        assert class_profile.expected_counts["inheritance"].expected_quantity == 1
        assert class_profile.mode == ScoringMode.EXPECTED_WITH_PENALTY

    def test_valores_de_secuencia_cuentan_por_fragmento(self, template_path):
        wb = load_workbook(template_path, data_only=True)
        profiles = parse_rubric_workbook(wb)
        seq_profile = profiles["sequence"]
        assert seq_profile.expected_counts["alt_fragments"].expected_quantity == 1
        assert seq_profile.expected_counts["loop_fragments"].expected_quantity == 1


class TestErroresDeFormato:
    def test_elemento_no_reconocido_acumula_error(self, template_path):
        wb = load_workbook(template_path, data_only=True)
        ws = wb["Clases"]
        ws.cell(row=2, column=1, value="ElementoInventado")
        with pytest.raises(RubricParseError) as exc_info:
            parse_rubric_workbook(wb)
        assert any("ElementoInventado" in e for e in exc_info.value.errors)

    def test_cantidad_no_numerica_acumula_error(self, template_path):
        wb = load_workbook(template_path, data_only=True)
        ws = wb["Clases"]
        ws.cell(row=2, column=2, value="cuatro")
        with pytest.raises(RubricParseError) as exc_info:
            parse_rubric_workbook(wb)
        assert any("Cantidad esperada" in e for e in exc_info.value.errors)

    def test_modo_no_reconocido_acumula_error(self, template_path):
        wb = load_workbook(template_path, data_only=True)
        ws = wb["Config"]
        ws.cell(row=2, column=2, value="modo_inventado")
        with pytest.raises(RubricParseError) as exc_info:
            parse_rubric_workbook(wb)
        assert any("modo_inventado" in e for e in exc_info.value.errors)


class TestModosRetirados:
    """Rúbricas ya distribuidas con 'hybrid' o 'range_tolerance' (o sus
    etiquetas en español ya retiradas) no deben rechazarse: se mapean al
    modo vigente más cercano."""

    def test_codigo_interno_hybrid_se_mapea_a_expected_with_penalty(self, template_path):
        wb = load_workbook(template_path, data_only=True)
        wb["Config"].cell(row=2, column=2, value="hybrid")
        profiles = parse_rubric_workbook(wb)
        assert profiles["class"].mode == ScoringMode.EXPECTED_WITH_PENALTY

    def test_codigo_interno_range_tolerance_se_mapea(self, template_path):
        wb = load_workbook(template_path, data_only=True)
        wb["Config"].cell(row=2, column=2, value="range_tolerance")
        profiles = parse_rubric_workbook(wb)
        assert profiles["class"].mode == ScoringMode.EXPECTED_WITH_PENALTY

    def test_etiqueta_espanol_retirada_se_mapea(self, template_path):
        wb = load_workbook(template_path, data_only=True)
        wb["Config"].cell(row=2, column=2, value="Híbrido (similitud + cantidades esperadas)")
        profiles = parse_rubric_workbook(wb)
        assert profiles["class"].mode == ScoringMode.EXPECTED_WITH_PENALTY
